#!/usr/bin/env python3
"""
Extract VT_Weekly_Template_FINAL.xlsx → data.json

Usage:
  LOCAL_XLSX_PATH=path/to/file.xlsx python update_data.py   # local
  (Azure env vars set)                python update_data.py   # SharePoint
"""
import io, json, math, os, re, sys
from datetime import date, datetime

import openpyxl

ZONES = ["Toronto", "Ottawa", "West", "East", "Northwest", "Essa", "Southwest", "Northeast", "Southeast"]
DAYS_OF_WEEK = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


# ── SharePoint download ───────────────────────────────────────────────────────
def fetch_from_sharepoint() -> bytes:
    import msal, requests

    tenant_id     = os.environ["AZURE_TENANT_ID"]
    client_id     = os.environ["AZURE_CLIENT_ID"]
    client_secret = os.environ["AZURE_CLIENT_SECRET"]
    site_hostname = os.environ.get("SP_SITE_HOSTNAME", "peakpowerenergy.sharepoint.com")
    site_path     = os.environ.get("SP_SITE_PATH",     "/sites/PeakPowerEnergy")
    file_name     = os.environ.get("SP_FILE_NAME",     "VT_Weekly_Template_FINAL.xlsx")

    app = msal.ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        client_credential=client_secret,
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Auth failed: {result.get('error_description')}")

    token   = result["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Resolve site ID
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_hostname}:{site_path}",
        headers=headers, timeout=30,
    )
    r.raise_for_status()
    site_id = r.json()["id"]

    # Search for file in drive root
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/search(q='{file_name}')",
        headers=headers, timeout=30,
    )
    r.raise_for_status()
    items  = r.json().get("value", [])
    target = next((i for i in items if i["name"] == file_name), None)
    if not target:
        raise FileNotFoundError(f"{file_name} not found in SharePoint drive")

    # Download
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{target['id']}/content",
        headers=headers, timeout=60,
    )
    r.raise_for_status()
    return r.content


# ── Helpers ──────────────────────────────────────────────────────────────────
def _date(v):
    if isinstance(v, datetime): return v.date().isoformat()
    if isinstance(v, date):     return v.isoformat()
    return str(v) if v else None

def _val(v, default=None):
    if v is None: return default
    if isinstance(v, float) and math.isnan(v): return default
    return v

def _float(v, default=0.0):
    v = _val(v)
    return float(v) if v is not None else default


# ── Extraction ───────────────────────────────────────────────────────────────
def extract_data(xlsx_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    # ── 1_Week Info ──────────────────────────────────────────────────────────
    weeks_raw = []
    for i, row in enumerate(wb["1_Week Info"].iter_rows(values_only=True)):
        if i < 3: continue
        if not isinstance(row[0], (int, float)): continue
        weeks_raw.append({
            "num":              int(row[0]),
            "start_date":       _date(row[1]),
            "end_date":         _date(row[2]),
            "phase":            _val(row[3], ""),
            "prepared_by":      _val(row[4], ""),
            "capital":          _float(row[5]),
            "starting_mwh":     _float(row[6]),
            "daily_risk_budget": _float(row[7]),
        })

    # ── 2_Daily Log ──────────────────────────────────────────────────────────
    daily_log_raw = []
    for i, row in enumerate(wb["2_Daily Log"].iter_rows(values_only=True)):
        if i < 3: continue
        if not isinstance(row[0], datetime): continue
        row = list(row) + [None] * max(0, 21 - len(row))
        daily_log_raw.append({
            "date":             _date(row[0]),
            "day":              _val(row[1], ""),
            "week_num":         int(_val(row[2], 0)),
            "day_type":         _val(row[3], ""),
            "t2_classified_by": _val(row[4], ""),
            "t1_on_duty":       _val(row[5], ""),
            "bid_submitted_by": _val(row[6], ""),
            "zones": {z: _val(row[7 + j]) for j, z in enumerate(ZONES)},
            "total_pnl":        _float(row[16]),
            "avg_spread":       _val(row[17]),
            "hit_rate":         _val(row[18]),
            "mwh_submitted":    _float(row[19]),
            "notes":            _val(row[20], ""),
        })

    # ── 3_Weekly Summary ─────────────────────────────────────────────────────
    SUMM_KEYS = {
        "Total Weekly P&L":    "total_pnl",
        "Trading Days":        "active_days",
        "Days Skipped":        "skipped_days",
        "Average Daily P&L":   "avg_daily_pnl",
        "Best Day P&L":        "best_day_pnl",
        "Worst Day P&L":       "worst_day_pnl",
        "Overall Hit Rate":    "hit_rate",
        "Total MWh":           "total_mwh",
        "Cumulative Losses":   "cumulative_losses",
        "Sharpe Ratio":        "sharpe_ratio",
    }
    summaries_raw = {}
    cur_wk = None
    for row in wb["3_Weekly Summary"].iter_rows(values_only=True):
        label = str(row[0] or "")
        m = re.search(r"WEEK\s+(\d+)", label.upper())
        if m:
            cur_wk = int(m.group(1))
            summaries_raw[cur_wk] = {}
            continue
        if cur_wk is None: continue
        for kw, key in SUMM_KEYS.items():
            if label.startswith(kw) and isinstance(row[1], (int, float)):
                summaries_raw[cur_wk][key] = row[1]
                break

    # ── 4_Zonal Matrix ───────────────────────────────────────────────────────
    zonal_matrices = {}
    cur_wk = None
    zone_idx = 0
    for row in wb["4_Zonal Matrix"].iter_rows(values_only=True):
        label = str(row[0] or "")
        m = re.search(r"WEEK\s+(\d+)", label.upper())
        if m:
            cur_wk = int(m.group(1))
            zonal_matrices[cur_wk] = {
                z: {d: 0.0 for d in DAYS_OF_WEEK} for z in ZONES
            }
            zonal_matrices[cur_wk]["_totals"]      = {z: 0.0 for z in ZONES}
            zonal_matrices[cur_wk]["_days_traded"]  = {z: 0   for z in ZONES}
            zone_idx = 0
            continue
        if cur_wk is None or zone_idx >= len(ZONES): continue
        zone = ZONES[zone_idx]
        for di, day in enumerate(DAYS_OF_WEEK):
            zonal_matrices[cur_wk][zone][day] = _float(row[1 + di])
        zonal_matrices[cur_wk]["_totals"][zone]      = _float(row[8] if len(row) > 8 else None)
        zonal_matrices[cur_wk]["_days_traded"][zone] = int(_float(row[9] if len(row) > 9 else None))
        zone_idx += 1

    # ── 5_Insights ───────────────────────────────────────────────────────────
    insights_raw = {}
    cur_wk = None
    cur_section = None
    for row in wb["5_Insights"].iter_rows(values_only=True):
        label = str(row[0] or row[1] or "")
        m = re.search(r"WEEK\s+(\d+)", label.upper())
        if m:
            cur_wk = int(m.group(1))
            insights_raw[cur_wk] = {"learnings": [], "issues": [], "action": {}}
            cur_section = None
            continue
        if cur_wk is None: continue
        if re.search(r"Section 3|Top Learn", label, re.I):   cur_section = "learnings"
        elif re.search(r"Section 4|Repeated", label, re.I):  cur_section = "issues"
        elif re.search(r"Section 5|Action",   label, re.I):  cur_section = "action"
        elif cur_section in ("learnings", "issues") and _val(row[1]):
            txt = str(row[1]).strip()
            if txt and not txt.startswith("#"):
                insights_raw[cur_wk][cur_section].append(txt)
        elif cur_section == "action" and _val(row[0]) and _val(row[1]):
            insights_raw[cur_wk]["action"][str(row[0])] = str(row[1])

    # ── Assemble per-week objects ─────────────────────────────────────────────
    weeks_out = []
    for w in weeks_raw:
        n      = w["num"]
        daily  = [d for d in daily_log_raw if d["week_num"] == n]
        active = [d for d in daily if d.get("day_type") != "Skipped"]

        summ = summaries_raw.get(n, {})
        if not summ:
            pnls  = [d["total_pnl"] for d in active]
            hrs   = [d["hit_rate"]  for d in active if d.get("hit_rate") is not None]
            losses = sum(abs(p) for p in pnls if p < 0)
            summ = {
                "total_pnl":        sum(pnls),
                "active_days":      len(active),
                "skipped_days":     len(daily) - len(active),
                "avg_daily_pnl":    sum(pnls) / len(pnls) if pnls else 0,
                "best_day_pnl":     max(pnls) if pnls else 0,
                "worst_day_pnl":    min(pnls) if pnls else 0,
                "hit_rate":         sum(hrs) / len(hrs) if hrs else 0,
                "total_mwh":        sum(d.get("mwh_submitted") or 0 for d in active),
                "cumulative_losses": losses,
                "sharpe_ratio":     None,
            }

        capital   = w.get("capital") or 0
        cum_loss  = abs(_float(summ.get("cumulative_losses")))
        cap_rem   = capital - cum_loss

        weeks_out.append({
            **w,
            "summary":          summ,
            "capital_remaining": cap_rem,
            "daily_log":        daily,
            "zonal_matrix":     zonal_matrices.get(n, {}),
            "insights":         insights_raw.get(n, {"learnings": [], "issues": [], "action": {}}),
        })

    # ── Aggregate ────────────────────────────────────────────────────────────
    all_active = [d for w in weeks_out for d in w["daily_log"] if d.get("day_type") != "Skipped"]
    all_pnls   = [d["total_pnl"] for d in all_active]
    all_hrs    = [d["hit_rate"]  for d in all_active if d.get("hit_rate") is not None]
    all_losses = sum(abs(p) for p in all_pnls if p < 0)

    zone_totals      = {}
    zone_days_traded = {}
    for z in ZONES:
        vals = [d["zones"].get(z) for d in all_active if d["zones"].get(z) is not None]
        zone_totals[z]      = sum(vals)
        zone_days_traded[z] = sum(1 for v in vals if v != 0)

    first_cap = weeks_out[0]["capital"] if weeks_out else 0
    aggregate = {
        "total_pnl":           sum(all_pnls),
        "active_days":         len(all_active),
        "avg_daily_pnl":       sum(all_pnls) / len(all_pnls) if all_pnls else 0,
        "best_day_pnl":        max(all_pnls) if all_pnls else 0,
        "worst_day_pnl":       min(all_pnls) if all_pnls else 0,
        "hit_rate":            sum(all_hrs) / len(all_hrs) if all_hrs else 0,
        "total_mwh":           sum(d.get("mwh_submitted") or 0 for d in all_active),
        "model_days":          sum(1 for d in all_active if d.get("day_type") == "Model Day"),
        "strategy_days":       sum(1 for d in all_active if d.get("day_type") == "Strategy Day"),
        "capital_deployed":    first_cap,
        "cumulative_losses":   all_losses,
        "capital_remaining":   first_cap - all_losses,
        "capital_preserved_pct": (first_cap - all_losses) / first_cap if first_cap else 0,
        "zone_totals":         zone_totals,
        "zone_days_traded":    zone_days_traded,
    }

    return {
        "last_updated":  datetime.utcnow().isoformat() + "Z",
        "source_file":   "VT_Weekly_Template_FINAL.xlsx",
        "zones":         ZONES,
        "days_of_week":  DAYS_OF_WEEK,
        "aggregate":     aggregate,
        "weeks":         weeks_out,
    }


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    local = os.environ.get("LOCAL_XLSX_PATH")
    if local and os.path.exists(local):
        print(f"Reading local file: {local}")
        with open(local, "rb") as f:
            xlsx_bytes = f.read()
    elif all(k in os.environ for k in ["AZURE_TENANT_ID", "AZURE_CLIENT_ID", "AZURE_CLIENT_SECRET"]):
        print("Fetching from SharePoint …")
        xlsx_bytes = fetch_from_sharepoint()
    else:
        sys.exit("ERROR: set LOCAL_XLSX_PATH or AZURE_TENANT_ID / CLIENT_ID / CLIENT_SECRET")

    print("Extracting data …")
    data     = extract_data(xlsx_bytes)
    out_path = os.environ.get("DATA_JSON_PATH", "data.json")
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2, default=str)

    wks = len(data["weeks"])
    days = data["aggregate"]["active_days"]
    print(f"✓ {out_path}  —  {wks} weeks · {days} active days")


if __name__ == "__main__":
    main()
